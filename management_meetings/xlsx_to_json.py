import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import openpyxl
import json

class ExcelToJsonConverter:
    """
    A GUI application for converting Excel sheets or tables to JSON format.

    This application uses tkinter for the GUI, pandas for reading Excel data,
    and openpyxl to inspect Excel files for sheet and table names.
    """
    def __init__(self, root):
        """
        Initializes the application's user interface.
        """
        self.root = root
        self.root.title("Excel to JSON Converter")
        self.root.geometry("550x350")
        self.root.resizable(False, False)
        
        # Style configuration
        style = ttk.Style(self.root)
        style.theme_use('clam')
        style.configure('TButton', padding=6, relief="flat", background="#0078D7", foreground="white")
        style.map('TButton', background=[('active', '#005A9E')])
        style.configure('TLabel', padding=5, font=('Helvetica', 10))
        style.configure('TFrame', padding=10)
        style.configure('TMenubutton', padding=6)

        # --- Member variables ---
        self.file_path = tk.StringVar()
        self.selected_sheet = tk.StringVar()
        self.selected_table = tk.StringVar()
        self.sheets = []
        self.tables = []

        # --- Main Frame ---
        main_frame = ttk.Frame(self.root, padding=(20, 10))
        main_frame.pack(fill=tk.BOTH, expand=True)

        # --- File Selection Frame ---
        file_frame = ttk.Frame(main_frame)
        file_frame.pack(fill=tk.X, pady=10)
        
        ttk.Label(file_frame, text="Excel File:").pack(side=tk.LEFT, padx=(0, 10))
        self.file_entry = ttk.Entry(file_frame, textvariable=self.file_path, state='readonly', width=50)
        self.file_entry.pack(side=tk.LEFT, expand=True, fill=tk.X)
        self.browse_button = ttk.Button(file_frame, text="Browse...", command=self.browse_file)
        self.browse_button.pack(side=tk.LEFT, padx=(10, 0))

        # --- Dropdown Selection Frame ---
        options_frame = ttk.Frame(main_frame)
        options_frame.pack(fill=tk.X, pady=10)

        # Sheet selection
        ttk.Label(options_frame, text="Select Sheet:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        self.sheet_menu = ttk.OptionMenu(options_frame, self.selected_sheet, "No file selected", *[""])
        self.sheet_menu.grid(row=0, column=1, sticky=tk.EW)
        self.selected_sheet.trace('w', self.on_sheet_select)

        # Table selection
        ttk.Label(options_frame, text="Select Table (Optional):").grid(row=1, column=0, sticky=tk.W, padx=(0, 10))
        self.table_menu = ttk.OptionMenu(options_frame, self.selected_table, "Select a sheet first", *[""])
        self.table_menu.grid(row=1, column=1, sticky=tk.EW)
        
        options_frame.grid_columnconfigure(1, weight=1)

        # --- Action Frame ---
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=20)

        self.convert_button = ttk.Button(action_frame, text="Convert to JSON", command=self.convert_to_json, state=tk.DISABLED)
        self.convert_button.pack(expand=True)

        # --- Status Bar ---
        self.status_var = tk.StringVar()
        self.status_var.set("Ready. Please select an Excel file to begin.")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=5)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def browse_file(self):
        """
        Opens a file dialog to select an .xlsx file and loads its sheets.
        """
        path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            return

        self.file_path.set(path)
        self.status_var.set(f"Loading sheets from: {path.split('/')[-1]}")
        self.root.update_idletasks() # Update GUI to show message

        try:
            # Use openpyxl to get sheet names without loading the whole file into memory
            workbook = openpyxl.load_workbook(path, read_only=True)
            self.sheets = workbook.sheetnames
            workbook.close()

            # Update sheet dropdown
            self.sheet_menu['menu'].delete(0, 'end')
            for sheet in self.sheets:
                self.sheet_menu['menu'].add_command(label=sheet, command=tk._setit(self.selected_sheet, sheet))
            
            if self.sheets:
                self.selected_sheet.set(self.sheets[0])
            else:
                self.selected_sheet.set("No sheets found")
                self.convert_button.config(state=tk.DISABLED)

            self.status_var.set("File loaded. Please select a sheet.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read Excel file.\nError: {e}")
            self.file_path.set("")
            self.status_var.set("Error loading file. Please try again.")

    def on_sheet_select(self, *args):
        """
        Callback for when a sheet is selected. It loads the tables for that sheet.
        """
        sheet_name = self.selected_sheet.get()
        if not self.file_path.get() or not sheet_name or sheet_name == "No sheets found":
            self.convert_button.config(state=tk.DISABLED)
            return

        try:
            # Load the workbook without read_only mode to access table information.
            # The 'tables' attribute is not available in read_only mode.
            workbook = openpyxl.load_workbook(self.file_path.get())
            sheet = workbook[sheet_name]
            
            # Get table names from the selected sheet
            self.tables = list(sheet.tables.keys())
            workbook.close()

            # Update table dropdown
            self.table_menu['menu'].delete(0, 'end')
            
            # The default option is to convert the entire sheet
            default_table_option = "Entire Sheet"
            self.tables.insert(0, default_table_option)
            
            for table in self.tables:
                self.table_menu['menu'].add_command(label=table, command=tk._setit(self.selected_table, table))
            
            self.selected_table.set(default_table_option)
            self.convert_button.config(state=tk.NORMAL) # Enable conversion button
            self.status_var.set("Ready to convert.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to read sheet '{sheet_name}'.\nError: {e}")
            self.convert_button.config(state=tk.DISABLED)
            self.status_var.set(f"Error reading sheet: {sheet_name}")

    def convert_to_json(self):
        """
        Performs the conversion from the selected Excel sheet/table to JSON.
        """
        file_path = self.file_path.get()
        sheet_name = self.selected_sheet.get()
        table_name = self.selected_table.get()

        if not file_path or not sheet_name:
            messagebox.showwarning("Warning", "Please select a file and a sheet first.")
            return

        output_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json")],
            initialfile=f"{sheet_name.replace(' ', '_')}_{table_name.replace(' ', '_') if table_name != 'Entire Sheet' else 'data'}.json"
        )

        if not output_path:
            return

        self.status_var.set("Converting... Please wait.")
        self.root.update_idletasks()

        try:
            json_data = ""
            if table_name == "Entire Sheet" or not table_name:
                # Use pandas to read the entire sheet
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Convert to JSON records format (list of objects)
                json_data = df.to_json(orient="records", indent=4, date_format='iso')
            else:
                # Use openpyxl to get the exact range of the table
                wb = openpyxl.load_workbook(file_path) # No read_only mode
                sheet = wb[sheet_name]
                table = sheet.tables[table_name]
                
                # Extract data from the table range
                data_rows = []
                cell_range = sheet[table.ref]
                for row in cell_range:
                    data_rows.append([cell.value for cell in row])
                
                wb.close()
                
                # First row is the header
                header = data_rows[0]
                records = data_rows[1:]
                
                # Create a list of dictionaries
                list_of_dicts = [dict(zip(header, row)) for row in records]
                
                # Convert to JSON string
                json_data = json.dumps(list_of_dicts, indent=4, default=str) # Use default=str for non-serializable types like dates

            # Write the JSON data to the output file
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(json_data)

            self.status_var.set("Conversion successful!")
            messagebox.showinfo("Success", f"File successfully converted and saved to:\n{output_path}")

        except Exception as e:
            self.status_var.set("Conversion failed!")
            messagebox.showerror("Error", f"An error occurred during conversion.\nError: {e}")


if __name__ == "__main__":
    # To run this script, you need to install the required libraries:
    # pip install pandas openpyxl
    
    root = tk.Tk()
    app = ExcelToJsonConverter(root)
    root.mainloop()

